"""
Excel-отчёт (openpyxl): лист исходных данных, лист параметров модели
и коэффициентов, лист результатов симуляции.
"""
from __future__ import annotations

import io

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def build_excel(state: dict, data) -> io.BytesIO:
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")

    def style_header(ws, row=1):
        for cell in ws[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    def autosize(ws):
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None),
                        default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = \
                min(width + 2, 40)

    # ------------------------------------------------ Лист 1: исходные данные
    ws = wb.active
    ws.title = "Исходные данные"
    ws.append(["Время, с", "PV", "SP", "CV, %"])
    for row in zip(data.time, data.pv, data.sp, data.cv):
        ws.append([round(float(v), 6) for v in row])
    style_header(ws); autosize(ws)

    # -------------------------------------- Лист 2: параметры и коэффициенты
    ws2 = wb.create_sheet("Параметры")
    ws2.append(["Параметр", "Значение"])
    rows = [
        ("Исходный файл", state.get("upload_name", "—")),
        ("K (коэффициент усиления)", round(state["K"], 6)),
        ("T (постоянная времени), с", round(state["T"], 4)),
        ("tau (запаздывание), с", round(state["tau"], 4)),
    ]
    if state.get("Ku"):
        rows.append(("Ku (критическое усиление)", round(state["Ku"], 5)))
        rows.append(("Tu (период автоколебаний), с", round(state["Tu"], 4)))
    coeffs = state["coeffs"]
    rows += [
        ("Kp регулятора", round(coeffs["Kp"], 6)),
        ("Ti, с", round(coeffs["Ti"], 4) if coeffs.get("Ti") else "—"),
        ("Td, с", round(coeffs["Td"], 4) if coeffs.get("Td") else "—"),
        ("Тип регулятора", state.get("ctype", "PID")),
        ("Метод настройки", state.get("tuning_method", "—")),
        ("Шаг дискретизации данных, с", round(data.dt, 6)),
    ]
    for r in rows:
        ws2.append(list(r))
    style_header(ws2); autosize(ws2)

    # --------------------------------------------------- Лист 3: симуляция
    ws3 = wb.create_sheet("Симуляция")
    sim = simulator_run(state, data)
    ws3.append(["Время, с", "SP", "PV (модель)", "CV, %"])
    for row in zip(*sim):
        ws3.append([round(float(v), 6) for v in row])
    style_header(ws3); autosize(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def simulator_run(state, data):
    """Запуск симуляции для листа отчёта."""
    from core import simulator
    coeffs = state["coeffs"]
    sim_time = min(max(2.0 * float(data.time[-1] - data.time[0]), 60.0), 3600.0)
    sim = simulator.simulate_closed_loop(
        state["K"], state["T"], state["tau"], state.get("ctype", "PID"),
        coeffs["Kp"], coeffs.get("Ti"), coeffs.get("Td"),
        dt_sim=max(data.dt / 5.0, 0.01), sim_time=sim_time,
        sp_array=data.sp)
    return sim
